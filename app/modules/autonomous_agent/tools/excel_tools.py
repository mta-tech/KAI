"""Excel file generation and reading tools."""
import os
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

def create_excel_tool(output_dir: str = "./agent_results"):
    """Create Excel writing tool."""

    def write_excel(
        data_json: str,
        filename: str | None = None,
        sheet_name: str = "Sheet1",
        table_style: str = "TableStyleMedium9"
    ) -> str:
        """Write data to an Excel file formatted as a table.

        Args:
            data_json: JSON string of data (list of dicts).
            filename: Optional filename (e.g. 'sales_data.xlsx'). If not provided, one will be generated.
            sheet_name: Name of the sheet to write data to.
            table_style: Excel table style (e.g., 'TableStyleMedium9', 'TableStyleLight1').

        Returns:
            JSON string with success status and file path.
        """
        try:
            # ensure output directory
            if not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)

            # parse data
            try:
                data = json.loads(data_json)
                # Handle wrapped data
                if isinstance(data, dict):
                    if "data" in data:
                        data = data["data"]
                    elif "result" in data:
                        data = data["result"]
            except json.JSONDecodeError:
                return json.dumps({"success": False, "error": "Invalid JSON data"})

            if not isinstance(data, list):
                return json.dumps({"success": False, "error": "Data must be a list of dictionaries"})

            df = pd.DataFrame(data)

            # generate filename if needed
            if not filename:
                from datetime import datetime
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"export_{timestamp}.xlsx"
            
            if not filename.endswith(".xlsx"):
                filename += ".xlsx"

            full_path = os.path.join(output_dir, filename)

            # Write to Excel
            df.to_excel(full_path, index=False, sheet_name=sheet_name)

            # Format as Table using openpyxl
            wb = load_workbook(full_path)
            ws = wb[sheet_name]

            # Define table range
            # If dataframe is empty, this might fail, but we checked data is a list.
            # Check if dataframe is empty
            if not df.empty:
                min_col = 1
                max_col = len(df.columns)
                min_row = 1
                max_row = len(df) + 1 # +1 for header

                tab = Table(displayName="Data", ref=f"A1:{get_column_letter(max_col)}{max_row}")

                # Add a default style with striped rows and banded columns
                style = TableStyleInfo(
                    name=table_style, 
                    showFirstColumn=False,
                    showLastColumn=False, 
                    showRowStripes=True, 
                    showColumnStripes=False
                )
                tab.tableStyleInfo = style
                ws.add_table(tab)

                # Adjust column widths
                for column_cells in ws.columns:
                    length = max(len(str(cell.value) or "") for cell in column_cells)
                    # Add a little extra padding
                    ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

            wb.save(full_path)

            return json.dumps({
                "success": True,
                "message": f"Excel file written successfully to {full_path}",
                "file_path": f"/results/{filename}",
                "real_path": full_path
            })

        except Exception as e:
            return json.dumps({"success": False, "error": str(e)})

    return write_excel


def create_read_excel_tool(base_dir: str = "./agent_results"):
    """Create Excel reading tool."""

    def read_excel(
        file_path: str,
        sheet_name: str | None = None,
        limit: int = 100,
        include_sheet_names: bool = False
    ) -> str:
        """Read data from an Excel file.

        Args:
            file_path: Path to the Excel file (.xlsx, .xls). Can be absolute, relative to base_dir, or a virtual path starting with /results/.
            sheet_name: Optional name of the sheet to read. If None, reads the first sheet.
            limit: Maximum number of rows to return (default 100).
            include_sheet_names: If True, returns list of all sheet names in the file.

        Returns:
            JSON string with data, column names, and metadata.
        """
        try:
            # Handle virtual paths from agent
            if file_path.startswith("/results/"):
                # Strip /results/ prefix and join with base_dir
                stripped = file_path.replace("/results/", "", 1).lstrip("/")
                real_path = os.path.join(base_dir, stripped)
            else:
                real_path = file_path

            # Fallback checks if path doesn't exist
            if not os.path.exists(real_path):
                # Try checking if it's directly in base_dir
                direct_in_base = os.path.join(base_dir, os.path.basename(file_path))
                if os.path.exists(direct_in_base):
                    real_path = direct_in_base
                elif os.path.exists(file_path): # Check as absolute/relative to CWD
                    real_path = file_path
                else:
                    return json.dumps({"success": False, "error": f"File not found: {file_path} (checked {real_path})"})

            # If just asking for sheet names
            if include_sheet_names:
                xls = pd.ExcelFile(real_path)
                return json.dumps({
                    "success": True, 
                    "sheet_names": xls.sheet_names,
                    "file_path": file_path
                })

            # Read data
            # pd.read_excel can take sheet_name=None to read all, but we usually want one for agents to process
            if sheet_name:
                df = pd.read_excel(real_path, sheet_name=sheet_name, nrows=limit)
            else:
                df = pd.read_excel(real_path, nrows=limit) # Reads first sheet by default

            # Replace NaN with None for valid JSON
            df = df.where(pd.notnull(df), None)
            
            # Convert to dict
            data = df.to_dict(orient="records")
            columns = list(df.columns)

            return json.dumps({
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name or "Sheet1",
                "row_count": len(data),
                "columns": columns,
                "data": data,
                "truncated": len(data) >= limit
            }, default=str) # Handle datetime serialization

        except Exception as e:
            return json.dumps({"success": False, "error": str(e)})

    return read_excel
